from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from io import BytesIO
from pydantic import ValidationError
import openpyxl
import os

import models
import schemas
import database

app = FastAPI()

# CORS middleware 
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Dependency for DB session
def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Create user
@app.post("/users", response_model=schemas.UserOut, status_code=201)
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    existing_user = db.query(models.User).filter(models.User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="User with this email already exists")

    new_user = models.User(
        first_name=user.first_name,
        last_name=user.last_name,
        email=user.email,
        phone=user.phone,
        pan=user.pan
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

# Get all users
@app.get("/users", response_model=list[schemas.UserOut])
def get_all_users(db: Session = Depends(get_db)):
    users = db.query(models.User).all()
    return users

# Update user
@app.put("/users/{user_id}", response_model=schemas.UserOut)
def update_user(user_id: int, updated_user: schemas.UserUpdate, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.first_name = updated_user.first_name
    user.last_name = updated_user.last_name
    user.email = updated_user.email
    user.phone = updated_user.phone
    user.pan = updated_user.pan

    db.commit()
    db.refresh(user)
    return user

# Delete user
@app.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    db.delete(user)
    db.commit()
    return

# Download Excel template
@app.get("/download-template", response_class=FileResponse)
def download_sample_excel():
    filepath = "sample_template.xlsx"
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Email", "Phone Number", "PAN Number"])
        ws.append(["Parichaye", "Grover", "parichaye@example.com", "1234567890", "ABCDE1234F"])
        wb.save(filepath)

    return FileResponse(path=filepath, filename="sample_template.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Upload Excel for bulk insert
@app.post("/upload-excel")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx format")

    try:
        contents = file.file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        errors = []
        users_to_add = []

        for idx, row in enumerate(rows, start=2):
            try:
                first_name, last_name, email, phone, pan = row
                user_data = schemas.UserCreate(
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=phone,
                    pan=pan
                )
                users_to_add.append(user_data)

            except (ValidationError, ValueError) as e:
                errors.append({"row": idx, "error": str(e)})

        if errors:
            return JSONResponse(status_code=400, content={"errors": errors})

        for user_data in users_to_add:
            user = models.User(**user_data.dict())
            db.add(user)

        db.commit()
        return {"message": f"{len(users_to_add)} users successfully added."}

    finally:
        file.file.close()
